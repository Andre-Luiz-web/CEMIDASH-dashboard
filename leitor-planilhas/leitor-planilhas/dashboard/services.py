from __future__ import annotations

from pathlib import Path
from threading import Lock
from typing import Any, Dict, Iterable, List, Optional, Tuple

from django.conf import settings
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


_CACHE_LOCK = Lock()
_CACHE: Dict[str, Any] = {"signature": None, "data": None}


def invalidate_dataset_cache() -> None:
    """Limpa o cache local para forçar uma nova leitura das planilhas."""
    with _CACHE_LOCK:
        _CACHE["signature"] = None
        _CACHE["data"] = None


def ensure_turmas_sincronizadas() -> List[str]:
    """
    Garante que todas as turmas encontradas nas planilhas existam na tabela `Turma`.
    Retorna a lista atual de turmas.
    """
    dataset = load_dataset()
    turmas_encontradas = dataset.get("turmas", [])
    if not turmas_encontradas:
        return turmas_encontradas

    from .models import Turma  # Import tardio para evitar dependência circular

    existentes = set(Turma.objects.values_list("nome", flat=True))
    turmas_encontradas_set = set(turmas_encontradas)

    novos = [turma for turma in turmas_encontradas if turma not in existentes]
    if novos:
        Turma.objects.bulk_create([Turma(nome=turma_nome) for turma_nome in novos], ignore_conflicts=True)

    remover = existentes - turmas_encontradas_set
    if remover:
        Turma.objects.filter(nome__in=remover).delete()

    return turmas_encontradas


def load_dataset() -> Dict[str, Any]:
    """
    Lê todas as planilhas disponíveis e devolve um dicionário com:
        - students: lista de resultados por estudante
        - turmas: nomes das turmas encontradas
        - arquivos: nomes das planilhas carregadas
        - question_bank: metadados agregados por questão
    Os resultados ficam em cache e somente são reconstruídos se algum arquivo mudar.
    """
    base_dir = Path(getattr(settings, "PLANILHAS_DIR", settings.BASE_DIR / "planilhas"))
    if not base_dir.exists():
        return {"students": [], "turmas": [], "arquivos": [], "question_bank": {}}

    signature = tuple(
        sorted((str(path), path.stat().st_mtime_ns) for path in base_dir.glob("*.xlsx"))
    )

    with _CACHE_LOCK:
        if _CACHE["signature"] == signature:
            return _CACHE["data"]

    dataset = _build_dataset(base_dir)

    with _CACHE_LOCK:
        _CACHE["signature"] = signature
        _CACHE["data"] = dataset

    return dataset


def _build_dataset(base_dir: Path) -> Dict[str, Any]:
    students: List[Dict[str, Any]] = []
    question_bank: Dict[str, Dict[str, Any]] = {}
    turmas: set[str] = set()
    arquivos: set[str] = set()

    for file_path in sorted(base_dir.glob("*.xlsx")):
        arquivos.add(file_path.name)
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_payload = _parse_sheet(file_path.name, sheet)
            if not sheet_payload:
                continue

            students.extend(sheet_payload["students"])
            turmas.update(sheet_payload["turmas"])
            _merge_question_bank(question_bank, sheet_payload["question_bank"])

    return {
        "students": students,
        "turmas": sorted(turmas),
        "arquivos": sorted(arquivos),
        "question_bank": question_bank,
    }


def _parse_sheet(file_name: str, sheet: Worksheet) -> Optional[Dict[str, Any]]:
    header_row = _get_row_values(sheet, 2)
    if not header_row:
        return None

    if not (_normalize_str(header_row[0]) == "nº" and _normalize_str(header_row[1]) == "cpf"):
        # Estrutura não segue o formato esperado (ex.: planilhas-resumo)
        return None

    gabarito_row = _get_row_values(sheet, 3)
    pesos_row = _get_row_values(sheet, 4)
    if not gabarito_row:
        return None

    question_columns = _identify_question_columns(header_row)
    if not question_columns:
        return None

    gabarito_por_questao = {
        question: _normalize_answer(gabarito_row[idx])
        for idx, question in question_columns
        if _normalize_answer(gabarito_row[idx]) != "##"
    }
    pesos_por_questao = {
        question: float(pesos_row[idx])
        for idx, question in question_columns
        if pesos_row and _is_number(pesos_row[idx]) and gabarito_por_questao.get(question) is not None
    }

    turma_label = sheet.cell(row=1, column=1).value
    turma_nome = _extract_turma_nome(sheet.title, turma_label)

    dados_estudantes: List[Dict[str, Any]] = []
    question_bank: Dict[str, Dict[str, Any]] = {
        question: {
            "gabarito": gabarito_por_questao.get(question),
            "peso": pesos_por_questao.get(question, 0.0),
            "total_respostas": 0,
            "total_acertos": 0,
        }
        for _, question in question_columns
        if gabarito_por_questao.get(question) is not None
    }

    primeira_linha_aluno = _find_first_student_row(sheet, start_row=5)
    if primeira_linha_aluno is None:
        return None

    max_score = sum(pesos_por_questao.values())

    for row in sheet.iter_rows(min_row=primeira_linha_aluno, values_only=True):
        if not row or all(cell in (None, "") for cell in row):
            continue

        numero_bruto = row[0]
        if not _is_student_number(numero_bruto):
            continue

        nome = _normalize_name(row[2])
        if not nome:
            continue
        if _normalize_str(nome) == "gabarito":
            continue

        cpf = _format_identifier(row[1])
        respostas = {}
        acertos = 0
        acertos_peso = 0.0
        total_questoes_validas = 0

        for idx, question in question_columns:
            resposta = _normalize_answer(row[idx])
            gabarito = gabarito_por_questao.get(question)
            peso = pesos_por_questao.get(question, 0.0)

            if gabarito is None:
                continue

            total_questoes_validas += 1
            if resposta:
                respostas[question] = resposta
                question_bank[question]["total_respostas"] += 1
                if resposta == gabarito:
                    acertos += 1
                    acertos_peso += peso
                    question_bank[question]["total_acertos"] += 1
            else:
                respostas[question] = None

        nota_planilha = row[3] if _is_number(row[3]) else None
        nota_calculada = nota_planilha if nota_planilha is not None else acertos_peso
        nota = round(float(nota_calculada), 2) if nota_calculada is not None else 0.0
        percentual_nota = round((nota / max_score * 100) if max_score else 0.0, 2)
        percentual_acertos = round(
            (acertos / total_questoes_validas * 100) if total_questoes_validas else 0.0,
            2,
        )

        dados_estudantes.append(
            {
                "arquivo": file_name,
                "sheet": sheet.title,
                "turma": turma_nome,
                "numero": int(numero_bruto) if _is_number(numero_bruto) else str(numero_bruto),
                "cpf": cpf,
                "nome": nome,
                "nota": nota,
                "max_nota": max_score,
                "percentual_nota": percentual_nota,
                "acertos": acertos,
                "total_questoes": total_questoes_validas,
                "percentual_acertos": percentual_acertos,
                "respostas": respostas,
                "gabarito": gabarito_por_questao,
            }
        )

    if not dados_estudantes:
        return None

    return {
        "students": dados_estudantes,
        "turmas": {turma_nome},
        "question_bank": question_bank,
    }


def _get_row_values(sheet: Worksheet, row_number: int) -> Optional[List[Any]]:
    row = sheet[row_number]
    if not row:
        return None
    return [cell.value for cell in row]


def _identify_question_columns(header_row: Iterable[Any]) -> List[Tuple[int, str]]:
    columns: List[Tuple[int, str]] = []
    for idx, value in enumerate(header_row):
        if idx <= 3:
            continue
        if value in (None, ""):
            continue

        question = str(value).strip()
        if question.endswith(".0"):
            question = question[:-2]
        question = question or str(idx - 3)
        columns.append((idx, question))
    return columns


def _merge_question_bank(
    global_bank: Dict[str, Dict[str, Any]],
    sheet_bank: Dict[str, Dict[str, Any]],
) -> None:
    for question, stats in sheet_bank.items():
        entry = global_bank.setdefault(
            question,
            {
                "gabarito": stats.get("gabarito"),
                "peso": stats.get("peso", 0.0),
                "total_respostas": 0,
                "total_acertos": 0,
            },
        )
        entry["total_respostas"] += stats.get("total_respostas", 0)
        entry["total_acertos"] += stats.get("total_acertos", 0)
        if entry["gabarito"] is None:
            entry["gabarito"] = stats.get("gabarito")
        if not entry["peso"]:
            entry["peso"] = stats.get("peso", 0.0)


def _find_first_student_row(sheet: Worksheet, start_row: int = 5) -> Optional[int]:
    for row_index in range(start_row, sheet.max_row + 1):
        row = _get_row_values(sheet, row_index)
        if not row:
            continue
        if _is_student_number(row[0]):
            return row_index
    return None


def _is_student_number(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    value_str = _normalize_str(value)
    return bool(value_str and value_str.isdigit())


def _normalize_answer(value: Any) -> Optional[str]:
    if value in (None, "", " "):
        return None
    text = str(value).strip().upper()
    return text or None


def _normalize_name(value: Any) -> str:
    if value in (None, "", " "):
        return ""
    return str(value).strip()


def _format_identifier(value: Any) -> str:
    if value in (None, "", " "):
        return ""
    if isinstance(value, float):
        return str(int(value))
    text = str(value).strip()
    return text


def _extract_turma_nome(sheet_title: str, raw_label: Any) -> str:
    if isinstance(raw_label, str) and raw_label.strip():
        label = raw_label.split("(")[0].strip()
        return label or sheet_title
    return sheet_title


def _normalize_str(value: Any) -> str:
    return str(value).strip().lower() if value not in (None, "") else ""


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)
